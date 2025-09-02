import openpyxl
from copy import copy
from openpyxl.utils import get_column_letter, column_index_from_string

def process_excel(input_file, output_file, col_func_map):
    """
    input_file: source Excel file
    output_file: output Excel file
    col_func_map: dict { column (index/name/letter): processing_function }
    """
    wb = openpyxl.load_workbook(input_file)
    ws = wb.active  # or wb["SheetName"]

    # Get headers to resolve names
    headers = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column + 1)]

    # Resolve all columns with their specific function
    resolved_cols = []
    for col, func in col_func_map.items():
        if isinstance(col, int):  # already index
            resolved_cols.append((col, func))
        elif isinstance(col, str):
            if col in headers:  # header name
                resolved_cols.append((headers.index(col) + 1, func))
            else:  # assume column letter
                resolved_cols.append((column_index_from_string(col), func))

    # Sort descending so insertions don’t shift later columns
    resolved_cols.sort(key=lambda x: x[0], reverse=True)

    for source_col, process_func in resolved_cols:
        target_col = source_col + 1
        target_letter = get_column_letter(target_col)

        # Insert new column
        ws.insert_cols(target_col)

        # Copy formatting row by row
        for row in range(1, ws.max_row + 1):
            source_cell = ws.cell(row=row, column=source_col)
            target_cell = ws.cell(row=row, column=target_col)

            if row == 1:  # header
                target_cell.value = f"{ws.cell(row=1, column=source_col).value}_Processed"

            if source_cell.has_style:
                target_cell.font = copy(source_cell.font)
                target_cell.border = copy(source_cell.border)
                target_cell.fill = copy(source_cell.fill)
                target_cell.number_format = copy(source_cell.number_format)
                target_cell.protection = copy(source_cell.protection)
                target_cell.alignment = copy(source_cell.alignment)

        # Process values
        max_len = len(ws.cell(row=1, column=target_col).value or "")
        for row in range(2, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=source_col).value
            processed = process_func(cell_value)
            ws.cell(row=row, column=target_col).value = processed
            if processed is not None:
                max_len = max(max_len, len(str(processed)))

        # Auto-fit width
        ws.column_dimensions[target_letter].width = max_len + 2

    # Save workbook
    wb.save(output_file)
    print(f"✅ Processing complete. Output saved as {output_file}")


# ---------------- Interactive Prompt ---------------- #

if __name__ == "__main__":
    input_file = input("📂 Enter the path to the input Excel file: ").strip()
    output_file = input("💾 Enter the path to save the output Excel file: ").strip()

    # Example processing functions
    def uppercase(val):
        return val.upper() if isinstance(val, str) else val

    def double(val):
        return val * 2 if isinstance(val, (int, float)) else val

    def reverse_str(val):
        return val[::-1] if isinstance(val, str) else val

    # Define which columns to process and how
    col_func_map = {
        3: uppercase,       # Process 3rd column
        "Amount": double,   # Process by header name
        "E": reverse_str    # Process by column letter
    }

    process_excel(input_file, output_file, col_func_map)
