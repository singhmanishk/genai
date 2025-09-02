import openpyxl
from copy import copy
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter, column_index_from_string

def process_excel(input_file, output_file, col_func_map, col_name_map=None, enable_word_wrap=False, auto_row_height=False):
    """
    input_file: source Excel file
    output_file: output Excel file
    col_func_map: dict { column (index/name/letter): processing_function }
    col_name_map: dict { column (same key as col_func_map): custom new column name }
    enable_word_wrap: True/False -> whether to enable word wrap in new column
    auto_row_height: True/False -> whether to auto-adjust row height if wrap is enabled
    """
    wb = openpyxl.load_workbook(input_file)
    ws = wb.active  # or wb["SheetName"]

    # Get headers to resolve names
    headers = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column + 1)]

    # Resolve all columns with their specific function
    resolved_cols = []
    for col, func in col_func_map.items():
        if isinstance(col, int):  # already index
            resolved_cols.append((col, func, col))
        elif isinstance(col, str):
            if col in headers:  # header name
                resolved_cols.append((headers.index(col) + 1, func, col))
            else:  # assume column letter
                resolved_cols.append((column_index_from_string(col), func, col))

    # Sort descending so insertions don’t shift later columns
    resolved_cols.sort(key=lambda x: x[0], reverse=True)

    for source_col, process_func, col_key in resolved_cols:
        target_col = source_col + 1
        target_letter = get_column_letter(target_col)

        # Insert new column
        ws.insert_cols(target_col)

        # Determine new column name
        custom_name = None
        if col_name_map and col_key in col_name_map:
            custom_name = col_name_map[col_key]

        original_header = ws.cell(row=1, column=source_col).value or "Column"
        new_header = custom_name if custom_name else f"{original_header}_Processed"

        # Copy formatting row by row
        for row in range(1, ws.max_row + 1):
            source_cell = ws.cell(row=row, column=source_col)
            target_cell = ws.cell(row=row, column=target_col)

            if row == 1:  # header row
                target_cell.value = new_header

            if source_cell.has_style:
                target_cell.font = copy(source_cell.font)
                target_cell.border = copy(source_cell.border)
                target_cell.fill = copy(source_cell.fill)
                target_cell.number_format = copy(source_cell.number_format)
                target_cell.protection = copy(source_cell.protection)
                target_cell.alignment = copy(source_cell.alignment)

            # Enable word wrap if requested
            if enable_word_wrap:
                target_cell.alignment = Alignment(
                    horizontal=target_cell.alignment.horizontal if target_cell.alignment else "general",
                    vertical=target_cell.alignment.vertical if target_cell.alignment else "bottom",
                    wrapText=True
                )

        # Process values
        max_len = len(new_header)
        for row in range(2, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=source_col).value
            processed = process_func(cell_value)
            ws.cell(row=row, column=target_col).value = processed
            if processed is not None:
                max_len = max(max_len, len(str(processed)))

        # Expand column width
        ws.column_dimensions[target_letter].width = max_len + 2

        # Auto-adjust row height if chosen
        if enable_word_wrap and auto_row_height:
            col_width = ws.column_dimensions[target_letter].width or 10
            for row in range(2, ws.max_row + 1):
                val = ws.cell(row=row, column=target_col).value
                if val:
                    text_len = len(str(val))
                    est_lines = max(1, int(text_len / col_width) + 1)
                    ws.row_dimensions[row].height = est_lines * 15  # ~15 points per line

    # Save workbook
    wb.save(output_file)
    print(f"✅ Processing complete. Output saved as {output_file}")


# ---------------- Interactive Prompt ---------------- #

if __name__ == "__main__":
    input_file = input("📂 Enter the path to the input Excel file: ").strip()
    output_file = input("💾 Enter the path to save the output Excel file: ").strip()
    
    wrap_choice = input("🔄 Enable word wrap in new column? (y/n): ").strip().lower()
    enable_word_wrap = wrap_choice == "y"
    
    auto_row_height = False
    if enable_word_wrap:
        row_choice = input("📏 Auto-adjust row height for wrapped text? (y/n): ").strip().lower()
        auto_row_height = row_choice == "y"

    # Example processing functions
    def uppercase(val):
        return val.upper() if isinstance(val, str) else val

    def double(val):
        return val * 2 if isinstance(val, (int, float)) else val

    def reverse_str(val):
        return val[::-1] if isinstance(val, str) else val

    # Define which columns to process
    col_func_map = {
        3: uppercase,       # Process 3rd column
        "Amount": double,   # Process by header name
        "E": reverse_str    # Process by column letter
    }

    # Ask user for custom names
    col_name_map = {}
    for col in col_func_map.keys():
        name = input(f"📝 Enter new column name for {col} (leave blank for default): ").strip()
        if name:
            col_name_map[col] = name

    process_excel(input_file, output_file, col_func_map, col_name_map, enable_word_wrap, auto_row_height)
